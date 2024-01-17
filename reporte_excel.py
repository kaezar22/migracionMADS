import os
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

def can_open_file(file_path):
    try:
        with open(file_path, 'rb'):
            pass
        return True
    except Exception as e:
        print(f"Error opening file {file_path}: {e}")
        return False

def timestamp_to_datetime(timestamp):
    return datetime.fromtimestamp(timestamp).strftime('%d/%m/%y %I:%M %p')

def create_excel_report(folder_path, report_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"
    ws.append(['Nombre de Archivo', 'Tipo', 'Tamaño (bytes)', 'Puede Abrirse', 'Última Modificación', 'Subcarpeta 1', 'Subcarpeta 2', 'Subcarpeta 3', 'Subcarpeta 4', 'Subcarpeta 5', 'Subcarpeta 6', 'Subcarpeta 7', 'Subcarpeta 8', 'Subcarpeta 9', 'Subcarpeta 10'])

    for root, dirs, files in os.walk(folder_path):
        # Exclude hidden folders
        dirs[:] = [d for d in dirs if not d.startswith('.')]

        for file_name in files:
            file_path = os.path.join(root, file_name)
            file_size = 0
            file_type = file_name.split('.')[-1] if '.' in file_name else 'Desconocido'
            can_open = can_open_file(file_path)

            try:
                file_size = os.path.getsize(file_path)
                last_modified = os.path.getmtime(file_path)
            except Exception as e:
                print(f"Error getting size or last modified for {file_path}: {e}")

            # Split the folder path into subfolders
            subfolders = os.path.relpath(root, folder_path).split(os.path.sep)[:10]

            # Fill in empty values if there are fewer than 10 subfolders
            subfolders.extend([''] * (10 - len(subfolders)))

            last_modified_formatted = timestamp_to_datetime(last_modified)

            ws.append([file_name, file_type, file_size, can_open, last_modified_formatted] + subfolders)

    # Auto-adjust column widths and apply alignment
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
        ws.column_dimensions[column].alignment = Alignment(horizontal='left', vertical='center')

    wb.save(report_path)

def main():
    # Set the folder path you want to check
    folder_path = r'D:\XXXXX\XXXXX' #ubicación y nombre de la carpeta a explorar

    # Set the path where you want to save the Excel report
    report_path = r'D:\XXXXXXXXXXXXXXXXXXXXX.xlsx'

    # Create Excel report
    create_excel_report(folder_path, report_path)

if __name__ == '__main__':
    main()