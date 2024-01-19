from office365.runtime.auth.authentication_context import AuthenticationContext
from office365.sharepoint.client_context import ClientContext
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

def timestamp_to_datetime(timestamp):
    return timestamp

def create_excel_report(ctx, folder, report_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "SharePoint Report"
    ws.append(['File Name', 'File Type', 'File Size (bytes)', 'Last Modification', 'Subfolder 1', 'Subfolder 2', 'Subfolder 3', 'Subfolder 4', 'Subfolder 5', 'Subfolder 6', 'Subfolder 7', 'Subfolder 8', 'Subfolder 9', 'Subfolder 10'])

    row_num = 2  # Start from the second row for data

    # List files and subfolders recursively
    row_num = list_files_and_subfolders(ctx, folder, ws, row_num)

    # Auto-adjust column widths and apply alignment
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
        ws.column_dimensions[column].alignment = Alignment(horizontal='left', vertical='center')

    wb.save(report_path)
    print(f"Excel report saved to {report_path}")

def list_files_and_subfolders(ctx, folder, sheet, row_num):
    # Get files in the current folder
    files = folder.files
    ctx.load(files)
    ctx.execute_query()

    # Extract and write information about each file to the Excel sheet
    for file in files:
        file_name = file.properties['Name']
        file_type = os.path.splitext(file_name)[1]  # Get file extension
        file_path = os.path.join(folder.properties['ServerRelativeUrl'], file_name)
        file_size = file.properties['Length']
        last_modified = file.properties['TimeLastModified']
        last_modified_formatted = timestamp_to_datetime(last_modified).strftime('%d/%m/%y %I:%M %p')

        # Split the folder path into subfolders
        subfolders = folder.properties['ServerRelativeUrl'].split('/')[4:14]  # Assumes the folder structure depth is up to 10 levels

        # Fill in empty values if there are fewer than 10 subfolders
        subfolders.extend([''] * (10 - len(subfolders)))

        sheet.append([file_name, file_type, file_size, last_modified_formatted] + subfolders)
        row_num += 1

    # Get subfolders in the current folder
    subfolders = folder.folders
    ctx.load(subfolders)
    ctx.execute_query()

    # Recursively process subfolders
    for subfolder in subfolders:
        row_num = list_files_and_subfolders(ctx, subfolder, sheet, row_num)

    return row_num

def main():
    # Your OneDrive for Business URL and credentials
    site_url = "https://ticminambiente-my.sharepoint.com/personal/ohdeviac_minambiente_gov_co"
    username = "usuarioXXXXXX"
    password = "contrasena"

    # Create an authentication context
    ctx_auth = AuthenticationContext(url=site_url)
    if ctx_auth.acquire_token_for_user(username, password):
        # Create a OneDrive for Business client context
        ctx = ClientContext(site_url, ctx_auth)

        # Specify the folder path within OneDrive
        folder_path = "/personal/ohdeviac_minambiente_gov_co/Documents/nombre_carpeta"

        # Get the root folder
        root_folder = ctx.web.get_folder_by_server_relative_url(folder_path)

        # Set the path where you want to save the Excel report
        report_path = r'D:\XXXXXXXXXX\SharePoint_carpeta_Report.xlsx'

        # Create Excel report
        create_excel_report(ctx, root_folder, report_path)

    else:
        print("Failed to authenticate.")

if __name__ == '__main__':
    main()